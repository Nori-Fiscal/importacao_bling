"""
excel_loader.py - Leitura de planilhas (.xlsx, .xls, .csv).
Suporta:
  - Planilha EAN  -> {sku: {"ean": "...", "descricao": "..."}}
  - Planilha XML F5 / espelho HAGN008 ->
      {sku: {"siscomex": Decimal, "afrmm": Decimal, "ibs": ..., "cbs": ...}}
"""

import csv
import io
import zipfile as zfcheck
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _limpar_digitos(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _dec(val) -> Decimal:
    if val is None:
        return Decimal("0")
    t = str(val).strip().replace(",", ".")
    if t.endswith(".0"):
        t = t[:-2]
    try:
        return Decimal(t)
    except InvalidOperation:
        return Decimal("0")


def ler_csv_tolerante(data: bytes) -> Tuple[List[List[str]], str]:
    text = data.decode("utf-8", errors="replace").replace("\x00", "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    try:
        dialect = csv.Sniffer().sniff(text[:5000])
        delim = dialect.delimiter
    except Exception:
        delim = ";"
    f = io.StringIO(text, newline="")
    reader = csv.reader(f, delimiter=delim, quotechar='"', doublequote=True)
    return list(reader), delim


def _find_col(header: List[str], nomes: List[str]) -> Optional[int]:
    for idx, h in enumerate(header):
        if h.strip().lower() in nomes:
            return idx + 1
    return None


def _normalizar_sku(s: str) -> str:
    return str(s or "").strip()


def _montar_registro_fiscal(
    siscomex,
    afrmm,
    base_ibs,
    aliq_ibs,
    valor_ibs,
    base_cbs,
    aliq_cbs,
    valor_cbs,
) -> Dict[str, Dict]:
    return {
        "siscomex": _dec(siscomex),
        "afrmm": _dec(afrmm),
        "ibs": {
            "base": _dec(base_ibs),
            "aliquota": _dec(aliq_ibs),
            "valor": _dec(valor_ibs),
        },
        "cbs": {
            "base": _dec(base_cbs),
            "aliquota": _dec(aliq_cbs),
            "valor": _dec(valor_cbs),
        },
    }


def carregar_planilha_ean(data: bytes, nome: str) -> Dict[str, Dict]:
    nome_lower = nome.lower()

    if zfcheck.is_zipfile(io.BytesIO(data)) or nome_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip().lower() for c in ws[1]]

        col_sku = (
            _find_col(header, ["sku", "codigo", "codigo (sku)", "cod"])
            or 3
        )
        col_ean = (
            _find_col(header, ["ean", "gtin", "codigo de barras", "barcode"])
            or 4
        )
        col_desc = _find_col(header, ["descricao", "desc", "produto", "xprod"])

        resultado: Dict[str, Dict] = {}
        for row in range(2, ws.max_row + 1):
            sku = _normalizar_sku(ws.cell(row=row, column=col_sku).value)
            if not sku:
                continue
            ean_raw = str(ws.cell(row=row, column=col_ean).value or "").strip()
            if ean_raw.endswith(".0"):
                ean_raw = ean_raw[:-2]
            ean = _limpar_digitos(ean_raw)
            desc = ""
            if col_desc:
                desc = str(ws.cell(row=row, column=col_desc).value or "").strip()
            if ean:
                resultado[sku] = {"ean": ean, "descricao": desc}
        return resultado

    if data[:4] == b"\xD0\xCF\x11\xE0" or nome_lower.endswith(".xls"):
        try:
            import xlrd
        except ImportError as e:
            raise ValueError("Arquivo .xls requer: pip install xlrd") from e
        book = xlrd.open_workbook(file_contents=data)
        ws = book.sheet_by_index(0)
        resultado: Dict[str, Dict] = {}
        for r in range(1, ws.nrows):
            sku = _normalizar_sku(ws.cell_value(r, 2))
            if not sku:
                continue
            ean_raw = str(ws.cell_value(r, 3)).strip()
            if ean_raw.endswith(".0"):
                ean_raw = ean_raw[:-2]
            ean = _limpar_digitos(ean_raw)
            if ean:
                resultado[sku] = {"ean": ean, "descricao": ""}
        return resultado

    rows, _ = ler_csv_tolerante(data)
    if not rows:
        return {}
    header = [h.strip().lower() for h in rows[0]]
    sku_i = None
    ean_i = None
    for i, h in enumerate(header):
        if h in ["sku", "codigo", "codigo (sku)"]:
            sku_i = i
        if h in ["ean", "gtin", "codigo de barras"]:
            ean_i = i
    start = 1 if (sku_i is not None and ean_i is not None) else 0
    sku_i = sku_i if sku_i is not None else 2
    ean_i = ean_i if ean_i is not None else 3

    resultado: Dict[str, Dict] = {}
    for r in rows[start:]:
        if len(r) <= max(sku_i, ean_i):
            continue
        sku = _normalizar_sku(r[sku_i])
        if not sku:
            continue
        ean_raw = str(r[ean_i]).strip()
        if ean_raw.endswith(".0"):
            ean_raw = ean_raw[:-2]
        ean = _limpar_digitos(ean_raw)
        if ean:
            resultado[sku] = {"ean": ean, "descricao": ""}
    return resultado


def carregar_planilha_xmlf5(data: bytes, nome: str) -> Dict[str, Dict]:
    nome_lower = nome.lower()

    if zfcheck.is_zipfile(io.BytesIO(data)) or nome_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip().lower() for c in ws[1]]

        col_sku = (
            _find_col(header, ["sku", "codigo", "codigo (sku)", "part number"])
            or 3
        )
        col_sis = _find_col(header, ["siscomex"]) or 4
        col_afr = _find_col(header, ["afrmm"]) or 5
        col_base_ibs = _find_col(header, ["base ibs"])
        col_aliq_ibs = _find_col(header, ["alq ibs"])
        col_valor_ibs = _find_col(header, ["valor ibs"])
        col_base_cbs = _find_col(header, ["base cbs"])
        col_aliq_cbs = _find_col(header, ["alq cbs"])
        col_valor_cbs = _find_col(header, ["valor cbs"])

        resultado: Dict[str, Dict] = {}
        for row in range(2, ws.max_row + 1):
            sku = _normalizar_sku(ws.cell(row=row, column=col_sku).value)
            if not sku or sku.lower().startswith("total"):
                continue
            resultado[sku] = _montar_registro_fiscal(
                ws.cell(row=row, column=col_sis).value if col_sis else None,
                ws.cell(row=row, column=col_afr).value if col_afr else None,
                ws.cell(row=row, column=col_base_ibs).value if col_base_ibs else None,
                ws.cell(row=row, column=col_aliq_ibs).value if col_aliq_ibs else None,
                ws.cell(row=row, column=col_valor_ibs).value if col_valor_ibs else None,
                ws.cell(row=row, column=col_base_cbs).value if col_base_cbs else None,
                ws.cell(row=row, column=col_aliq_cbs).value if col_aliq_cbs else None,
                ws.cell(row=row, column=col_valor_cbs).value if col_valor_cbs else None,
            )
        return resultado

    if data[:4] == b"\xD0\xCF\x11\xE0" or nome_lower.endswith(".xls"):
        try:
            import xlrd
        except ImportError as e:
            raise ValueError("Arquivo .xls requer: pip install xlrd") from e
        book = xlrd.open_workbook(file_contents=data)
        ws = book.sheet_by_index(0)
        header = [str(ws.cell_value(0, c) or "").strip().lower() for c in range(ws.ncols)]

        col_sku = next((i for i, h in enumerate(header) if h in ["sku", "codigo", "codigo (sku)", "part number"]), 2)
        col_sis = next((i for i, h in enumerate(header) if h == "siscomex"), 3)
        col_afr = next((i for i, h in enumerate(header) if h == "afrmm"), 4)
        col_base_ibs = next((i for i, h in enumerate(header) if h == "base ibs"), None)
        col_aliq_ibs = next((i for i, h in enumerate(header) if h == "alq ibs"), None)
        col_valor_ibs = next((i for i, h in enumerate(header) if h == "valor ibs"), None)
        col_base_cbs = next((i for i, h in enumerate(header) if h == "base cbs"), None)
        col_aliq_cbs = next((i for i, h in enumerate(header) if h == "alq cbs"), None)
        col_valor_cbs = next((i for i, h in enumerate(header) if h == "valor cbs"), None)

        resultado: Dict[str, Dict] = {}
        for r in range(1, ws.nrows):
            sku = _normalizar_sku(ws.cell_value(r, col_sku))
            if not sku or sku.lower().startswith("total"):
                continue
            resultado[sku] = _montar_registro_fiscal(
                ws.cell_value(r, col_sis) if col_sis is not None else None,
                ws.cell_value(r, col_afr) if col_afr is not None else None,
                ws.cell_value(r, col_base_ibs) if col_base_ibs is not None else None,
                ws.cell_value(r, col_aliq_ibs) if col_aliq_ibs is not None else None,
                ws.cell_value(r, col_valor_ibs) if col_valor_ibs is not None else None,
                ws.cell_value(r, col_base_cbs) if col_base_cbs is not None else None,
                ws.cell_value(r, col_aliq_cbs) if col_aliq_cbs is not None else None,
                ws.cell_value(r, col_valor_cbs) if col_valor_cbs is not None else None,
            )
        return resultado

    rows, _ = ler_csv_tolerante(data)
    if not rows:
        return {}
    header = [h.strip().lower() for h in rows[0]]
    sku_i = None
    sis_i = None
    afr_i = None
    for i, h in enumerate(header):
        if h in ["sku", "codigo", "codigo (sku)", "part number"]:
            sku_i = i
        if h == "siscomex":
            sis_i = i
        if h == "afrmm":
            afr_i = i
    base_ibs_i = next((i for i, h in enumerate(header) if h == "base ibs"), None)
    aliq_ibs_i = next((i for i, h in enumerate(header) if h == "alq ibs"), None)
    valor_ibs_i = next((i for i, h in enumerate(header) if h == "valor ibs"), None)
    base_cbs_i = next((i for i, h in enumerate(header) if h == "base cbs"), None)
    aliq_cbs_i = next((i for i, h in enumerate(header) if h == "alq cbs"), None)
    valor_cbs_i = next((i for i, h in enumerate(header) if h == "valor cbs"), None)
    start = 1 if (sku_i is not None) else 0
    sku_i = sku_i if sku_i is not None else 2
    sis_i = sis_i if sis_i is not None else 3
    afr_i = afr_i if afr_i is not None else 4

    resultado: Dict[str, Dict] = {}
    for r in rows[start:]:
        if len(r) <= max(sku_i, sis_i, afr_i):
            continue
        sku = _normalizar_sku(r[sku_i])
        if not sku or sku.lower().startswith("total"):
            continue
        resultado[sku] = _montar_registro_fiscal(
            r[sis_i] if sis_i is not None and len(r) > sis_i else None,
            r[afr_i] if afr_i is not None and len(r) > afr_i else None,
            r[base_ibs_i] if base_ibs_i is not None and len(r) > base_ibs_i else None,
            r[aliq_ibs_i] if aliq_ibs_i is not None and len(r) > aliq_ibs_i else None,
            r[valor_ibs_i] if valor_ibs_i is not None and len(r) > valor_ibs_i else None,
            r[base_cbs_i] if base_cbs_i is not None and len(r) > base_cbs_i else None,
            r[aliq_cbs_i] if aliq_cbs_i is not None and len(r) > aliq_cbs_i else None,
            r[valor_cbs_i] if valor_cbs_i is not None and len(r) > valor_cbs_i else None,
        )
    return resultado
