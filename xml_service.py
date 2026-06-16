"""
xml_service.py — Processamento de XMLs NF-e (importação DUIMP/F5).

Regras aplicadas:
  1. EAN: aplica da base de dados; se ausente, preserva o existente no XML
  2. cFabricante: igual a cProd (SKU) dentro de <adi>
  3. ICMS: zera todos os valores monetários e limpa infAdic/infCpl
  4. IPI CST → 03 (quando sem alíquota)
  5. PIS/COFINS CST → 50
"""

import io
import re
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Callable, Dict, List, Optional, Tuple

from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def _ns(tag: str) -> Optional[str]:
    if tag.startswith("{") and "}" in tag:
        return tag[1:].split("}")[0]
    return None


def _first(xpath_result: List) -> Optional[str]:
    if not xpath_result:
        return None
    v = xpath_result[0]
    return str(v).strip() if v is not None else None


def _dec(text) -> Decimal:
    if text is None:
        return Decimal("0")
    t = str(text).strip().replace(",", ".")
    if not t:
        return Decimal("0")
    try:
        return Decimal(t)
    except InvalidOperation:
        return Decimal("0")


def _q2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _limpar_digitos(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _mk(ns_uri: Optional[str], local: str, text: str) -> etree._Element:
    tag = f"{{{ns_uri}}}{local}" if ns_uri else local
    e = etree.Element(tag)
    e.text = text
    return e


def _get_text(elem: etree._Element, local: str) -> Optional[str]:
    r = elem.xpath(f"./*[local-name()='{local}']/text()")
    return str(r[0]).strip() if r else None


def _novo_stats() -> Dict:
    return {
        # EAN
        "ean_criados": 0,
        "ean_atualizados": 0,
        "ean_preservados": 0,     # sem base, mas XML já tinha valor
        "ean_ausentes": 0,        # sem base e sem valor no XML
        "faltantes_detalhado": [],

        # cFabricante
        "cfabricante_inseridos": 0,
        "cfabricante_atualizados": 0,
        "cfabricante_reposicionados": 0,
        "cfabricante_deduplicados": 0,

        # ICMS
        "icms_zerados": 0,
        "cfop_ajustados": 0,
        "infcpl_limpa": 0,

        # IPI / PIS / COFINS
        "ipi_cst_alterados": 0,
        "ipi_cst_criados": 0,
        "pis_cst_alterados": 0,
        "cofins_cst_alterados": 0,

        # IBS / CBS
        "ibscbs_itens_gerados": 0,
        "ibscbs_totais_gerados": 0,

        # CAMEX / Gecex
        "camex_alertas": 0,
        "camex_itens_detalhado": [],

        # TTD409 / Decreto SC 2.128/2009
        "ttd409_bloqueios": 0,
        "ttd409_itens_detalhado": [],

        "avisos": [],
        "erros": [],
    }


# ---------------------------------------------------------------------------
# 1. EAN — aplica da base; preserva se ausente (nunca destrói)
# ---------------------------------------------------------------------------

def _aplicar_ean(root: etree._Element, stats: Dict, mapa_ean: Dict[str, str], nome_arquivo: str) -> None:
    for det in root.xpath(".//*[local-name()='det']"):
        prod_nodes = det.xpath("./*[local-name()='prod']")
        if not prod_nodes:
            continue
        prod = prod_nodes[0]

        nitem = det.get("nItem", "")
        cprod_nodes = prod.xpath("./*[local-name()='cProd']")
        if not cprod_nodes:
            continue
        sku = (cprod_nodes[0].text or "").strip()
        if not sku:
            continue

        xprod = _first(prod.xpath("./*[local-name()='xProd']/text()")) or ""
        cean_orig = _first(prod.xpath("./*[local-name()='cEAN']/text()")) or ""
        ceantrib_orig = _first(prod.xpath("./*[local-name()='cEANTrib']/text()")) or ""

        gtin = _limpar_digitos(mapa_ean.get(sku, "") or "")

        ns_uri = _ns(prod.tag)
        tag_cean = f"{{{ns_uri}}}cEAN" if ns_uri else "cEAN"
        tag_ceantrib = f"{{{ns_uri}}}cEANTrib" if ns_uri else "cEANTrib"

        cean_nodes = prod.xpath("./*[local-name()='cEAN']")
        ceantrib_nodes = prod.xpath("./*[local-name()='cEANTrib']")
        idx_cprod = prod.index(cprod_nodes[0])

        if not gtin:
            # Sem EAN na base: PRESERVAR o que há no XML, apenas sinalizar
            stats["ean_ausentes"] += 1
            if cean_orig or ceantrib_orig:
                stats["ean_preservados"] += 1

            stats["faltantes_detalhado"].append({
                "Arquivo XML": nome_arquivo,
                "nItem": nitem,
                "SKU": sku,
                "Descrição": xprod,
                "cEAN no XML": cean_orig,
                "cEANTrib no XML": ceantrib_orig,
            })
            continue

        # Atualiza ou cria cEAN
        if cean_nodes:
            if (cean_nodes[0].text or "").strip() != gtin:
                cean_nodes[0].text = gtin
                stats["ean_atualizados"] += 1
        else:
            new_cean = etree.Element(tag_cean)
            new_cean.text = gtin
            prod.insert(idx_cprod + 1, new_cean)
            stats["ean_criados"] += 1

        # Atualiza ou cria cEANTrib
        cean_nodes = prod.xpath("./*[local-name()='cEAN']")
        if ceantrib_nodes:
            if (ceantrib_nodes[0].text or "").strip() != gtin:
                ceantrib_nodes[0].text = gtin
        else:
            new_t = etree.Element(tag_ceantrib)
            new_t.text = gtin
            ref = cean_nodes[0] if cean_nodes else None
            if ref is not None:
                prod.insert(prod.index(ref) + 1, new_t)
            else:
                prod.insert(idx_cprod + 2, new_t)


# ---------------------------------------------------------------------------
# 2. cFabricante = cProd dentro de <adi>
# ---------------------------------------------------------------------------

def _inserir_cfabricante(root: etree._Element, stats: Dict) -> None:
    for adi in root.xpath(".//*[local-name()='adi']"):
        det_anc = adi.xpath("ancestor::*[local-name()='det'][1]")
        if not det_anc:
            continue
        det = det_anc[0]
        cprod = det.xpath("string(./*[local-name()='prod']/*[local-name()='cProd'])").strip()
        if not cprod:
            stats["avisos"].append("adi encontrado, mas det não possui cProd.")
            continue

        ns_uri = _ns(adi.tag)
        tag_cfab = f"{{{ns_uri}}}cFabricante" if ns_uri else "cFabricante"

        nseq = next(iter(adi.xpath("./*[local-name()='nSeqAdic']")), None)

        existentes = [ch for ch in adi if isinstance(ch.tag, str) and etree.QName(ch).localname == "cFabricante"]
        if len(existentes) > 1:
            for extra in existentes[1:]:
                adi.remove(extra)
            existentes = existentes[:1]
            stats["cfabricante_deduplicados"] += 1

        if existentes:
            cf = existentes[0]
            if (cf.text or "").strip() != cprod:
                cf.text = cprod
                stats["cfabricante_atualizados"] += 1
        else:
            cf = etree.Element(tag_cfab)
            cf.text = cprod
            stats["cfabricante_inseridos"] += 1

        # Posicionar após nSeqAdic (ou no início)
        if nseq is not None:
            idx_nseq = adi.index(nseq)
            if cf in adi:
                if adi.index(cf) != idx_nseq + 1:
                    adi.remove(cf)
                    adi.insert(idx_nseq + 1, cf)
                    stats["cfabricante_reposicionados"] += 1
            else:
                adi.insert(idx_nseq + 1, cf)
        else:
            if cf in adi:
                if adi.index(cf) != 0:
                    adi.remove(cf)
                    adi.insert(0, cf)
                    stats["cfabricante_reposicionados"] += 1
            else:
                adi.insert(0, cf)


# ---------------------------------------------------------------------------
# 3. ICMS — zera todos os valores monetários; limpa infCpl
# ---------------------------------------------------------------------------

def _ajustar_cfop(root: etree._Element, stats: Dict) -> None:
    for cfop in root.xpath(".//*[local-name()='det']/*[local-name()='prod']/*[local-name()='CFOP']"):
        if (cfop.text or "").strip() == "3949":
            cfop.text = "3102"
            stats["cfop_ajustados"] += 1


_CAMPOS_MONETARIOS_ICMS = [
    "vBC", "vICMS", "vICMSOp", "vICMSDif",
    "vBCSTRet", "vICMSSTRet", "vICMSST",
    "vBCST", "vBCFCPST",
    "vFCPST", "vFCP", "vFCPSTRet",
    "pDif",
]

_ICMS_REGEX = re.compile(
    r"[^;.]*\bICMS\b[^;.]*[;.]?",
    flags=re.IGNORECASE,
)


def _zerar_icms(root: etree._Element, stats: Dict) -> None:
    """
    Para cada item det:
      - Garante que o grupo ICMS seja ICMS51 com CST=51
      - Zera todos os campos monetários (vBC, vICMS, etc.)
      - Preserva orig, modBC, pICMS
      - Remove referências a ICMS no infCpl
    """
    for det in root.xpath(".//*[local-name()='det']"):
        icms_wrapper = det.xpath(".//*[local-name()='imposto']//*[local-name()='ICMS']")
        if not icms_wrapper:
            continue
        icms = icms_wrapper[0]

        grupos = [ch for ch in icms if isinstance(ch.tag, str)]
        if not grupos:
            continue
        old = grupos[0]
        old_name = etree.QName(old).localname

        # Simples Nacional: não mexe
        if old_name.startswith("ICMSSN"):
            stats["avisos"].append(f"ICMS Simples Nacional ({old_name}) — não alterado.")
            continue

        ns_uri = _ns(old.tag) or _ns(icms.tag)

        # Lê campos que devem ser preservados
        orig   = _get_text(old, "orig")   or "0"
        modBC  = _get_text(old, "modBC")  or "3"
        pICMS  = _get_text(old, "pICMS")  or "0.00"
        pRedBC = _get_text(old, "pRedBC")

        # Monta novo nó ICMS51 com valores zerados
        tag_icms51 = f"{{{ns_uri}}}ICMS51" if ns_uri else "ICMS51"
        novo = etree.Element(tag_icms51)
        novo.append(_mk(ns_uri, "orig",      orig))
        novo.append(_mk(ns_uri, "CST",       "51"))
        novo.append(_mk(ns_uri, "modBC",     modBC))
        if pRedBC is not None:
            novo.append(_mk(ns_uri, "pRedBC", pRedBC))
        novo.append(_mk(ns_uri, "vBC",       "0.00"))
        novo.append(_mk(ns_uri, "pICMS",     "0.00"))
        novo.append(_mk(ns_uri, "vICMSOp",   "0.00"))
        novo.append(_mk(ns_uri, "pDif",      "0.00"))
        novo.append(_mk(ns_uri, "vICMSDif",  "0.00"))
        novo.append(_mk(ns_uri, "vICMS",     "0.00"))

        idx = icms.index(old)
        icms.remove(old)
        icms.insert(idx, novo)
        stats["icms_zerados"] += 1

    # Limpa infCpl de referências a ICMS
    for infcpl in root.xpath(".//*[local-name()='infCpl']"):
        if infcpl.text:
            original = infcpl.text
            limpo = _ICMS_REGEX.sub("", original)
            limpo = re.sub(r"\s{2,}", " ", limpo).strip()
            if limpo != original:
                infcpl.text = limpo
                stats["infcpl_limpa"] += 1

    icmstot = next(iter(root.xpath(".//*[local-name()='total']/*[local-name()='ICMSTot']")), None)
    if icmstot is not None:
        for campo in ["vBC", "vICMS"]:
            no = next(iter(icmstot.xpath(f"./*[local-name()='{campo}']")), None)
            if no is not None:
                no.text = "0.00"


# ---------------------------------------------------------------------------
# 4. IPI — CST → 03 quando sem alíquota
# ---------------------------------------------------------------------------

def _ajustar_ipi(root: etree._Element, stats: Dict) -> None:
    for ipi in root.xpath(".//*[local-name()='IPI']"):
        p_ipi = _first(ipi.xpath(".//*[local-name()='pIPI']/text()"))
        aliquota = _dec(p_ipi)

        if aliquota == 0:
            cst_nodes = ipi.xpath(".//*[local-name()='CST']")
            if cst_nodes:
                for cst in cst_nodes:
                    if (cst.text or "").strip() != "03":
                        cst.text = "03"
                        stats["ipi_cst_alterados"] += 1
            else:
                alvo = next(iter(
                    ipi.xpath("./*[local-name()='IPITrib']") or
                    ipi.xpath("./*[local-name()='IPINT']")
                ), None)
                if alvo is None:
                    stats["avisos"].append("IPI sem CST e sem IPITrib/IPINT: não foi possível inserir CST=03.")
                    continue
                ns_uri = _ns(alvo.tag)
                tag_cst = f"{{{ns_uri}}}CST" if ns_uri else "CST"
                cst_new = etree.Element(tag_cst)
                cst_new.text = "03"
                alvo.insert(0, cst_new)
                stats["ipi_cst_criados"] += 1


# ---------------------------------------------------------------------------
# 5. PIS / COFINS — CST → 50
# ---------------------------------------------------------------------------

def _ajustar_pis_cofins(root: etree._Element, stats: Dict) -> None:
    for pis in root.xpath(".//*[local-name()='PIS']"):
        for cst in pis.xpath(".//*[local-name()='CST']"):
            if (cst.text or "").strip() != "50":
                cst.text = "50"
                stats["pis_cst_alterados"] += 1

    for cofins in root.xpath(".//*[local-name()='COFINS']"):
        for cst in cofins.xpath(".//*[local-name()='CST']"):
            if (cst.text or "").strip() != "50":
                cst.text = "50"
                stats["cofins_cst_alterados"] += 1


# ---------------------------------------------------------------------------
# 6. IBS / CBS - inclui grupos conforme espelho HAGN008
# ---------------------------------------------------------------------------

_IBS_CBS_CFG = {
    "cst": "000",
    "cclass_trib": "000001",
    "p_ibs_uf": Decimal("0.1000"),
    "p_ibs_mun": Decimal("0.0000"),
    "p_cbs": Decimal("0.9000"),
}


def _fmt_q2(valor: Decimal) -> str:
    return f"{_q2(valor):.2f}"


def _fmt_q4(valor: Decimal) -> str:
    return f"{valor.quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP):.4f}"


def _append_or_replace(parent: etree._Element, child: etree._Element, local_name: str) -> None:
    existentes = [
        no for no in parent
        if isinstance(no.tag, str) and etree.QName(no).localname == local_name
    ]
    if existentes:
        idx = parent.index(existentes[0])
        for extra in existentes:
            parent.remove(extra)
        parent.insert(idx, child)
    else:
        parent.append(child)


def _adicionar_ibscbs(root: etree._Element, stats: Dict, mapa_fiscal: Optional[Dict[str, Dict]] = None) -> None:
    total_vbc = Decimal("0")
    total_vibsuf = Decimal("0")
    total_vibsmun = Decimal("0")
    total_vibs = Decimal("0")
    total_vcbs = Decimal("0")
    mapa_fiscal = mapa_fiscal or {}

    for det in root.xpath(".//*[local-name()='det']"):
        imposto = next(iter(det.xpath("./*[local-name()='imposto']")), None)
        prod = next(iter(det.xpath("./*[local-name()='prod']")), None)
        if imposto is None or prod is None:
            continue

        ns_uri = _ns(imposto.tag) or _ns(det.tag)
        sku = (_get_text(prod, "cProd") or "").strip()
        fiscal = mapa_fiscal.get(sku, {})
        ibs_cfg = fiscal.get("ibs", {}) if isinstance(fiscal, dict) else {}
        cbs_cfg = fiscal.get("cbs", {}) if isinstance(fiscal, dict) else {}

        vbc = _q2(
            _dec(ibs_cfg.get("base"))
            or _dec(cbs_cfg.get("base"))
            or _dec(_get_text(prod, "vProd"))
        )
        p_ibs_uf = _dec(ibs_cfg.get("aliquota")) or _IBS_CBS_CFG["p_ibs_uf"]
        p_ibs_mun = _IBS_CBS_CFG["p_ibs_mun"]
        p_cbs = _dec(cbs_cfg.get("aliquota")) or _IBS_CBS_CFG["p_cbs"]

        vibsuf = _q2(_dec(ibs_cfg.get("valor")) or (vbc * (p_ibs_uf / Decimal("100"))))
        vibsmun = _q2(vbc * (p_ibs_mun / Decimal("100")))
        vibs = _q2(vibsuf + vibsmun)
        vcbs = _q2(_dec(cbs_cfg.get("valor")) or (vbc * (p_cbs / Decimal("100"))))

        ibscbs = etree.Element(f"{{{ns_uri}}}IBSCBS" if ns_uri else "IBSCBS")
        ibscbs.append(_mk(ns_uri, "CST", _IBS_CBS_CFG["cst"]))
        ibscbs.append(_mk(ns_uri, "cClassTrib", _IBS_CBS_CFG["cclass_trib"]))

        gibscbs = etree.Element(f"{{{ns_uri}}}gIBSCBS" if ns_uri else "gIBSCBS")
        gibscbs.append(_mk(ns_uri, "vBC", _fmt_q2(vbc)))

        gibsuf = etree.Element(f"{{{ns_uri}}}gIBSUF" if ns_uri else "gIBSUF")
        gibsuf.append(_mk(ns_uri, "pIBSUF", _fmt_q4(p_ibs_uf)))
        gibsuf.append(_mk(ns_uri, "vIBSUF", _fmt_q2(vibsuf)))
        gibscbs.append(gibsuf)

        gibsmun = etree.Element(f"{{{ns_uri}}}gIBSMun" if ns_uri else "gIBSMun")
        gibsmun.append(_mk(ns_uri, "pIBSMun", _fmt_q4(p_ibs_mun)))
        gibsmun.append(_mk(ns_uri, "vIBSMun", _fmt_q2(vibsmun)))
        gibscbs.append(gibsmun)

        gibscbs.append(_mk(ns_uri, "vIBS", _fmt_q2(vibs)))

        gcbs = etree.Element(f"{{{ns_uri}}}gCBS" if ns_uri else "gCBS")
        gcbs.append(_mk(ns_uri, "pCBS", _fmt_q4(p_cbs)))
        gcbs.append(_mk(ns_uri, "vCBS", _fmt_q2(vcbs)))
        gibscbs.append(gcbs)

        ibscbs.append(gibscbs)
        _append_or_replace(imposto, ibscbs, "IBSCBS")

        total_vbc += vbc
        total_vibsuf += vibsuf
        total_vibsmun += vibsmun
        total_vibs += vibs
        total_vcbs += vcbs
        stats["ibscbs_itens_gerados"] += 1

    total = next(iter(root.xpath(".//*[local-name()='total']")), None)
    icmstot = next(iter(root.xpath(".//*[local-name()='total']/*[local-name()='ICMSTot']")), None)
    if total is None or icmstot is None:
        return

    ns_uri = _ns(total.tag) or _ns(icmstot.tag)
    ibscbstot = etree.Element(f"{{{ns_uri}}}IBSCBSTot" if ns_uri else "IBSCBSTot")
    ibscbstot.append(_mk(ns_uri, "vBCIBSCBS", _fmt_q2(total_vbc)))

    gibs = etree.Element(f"{{{ns_uri}}}gIBS" if ns_uri else "gIBS")
    gibsuf = etree.Element(f"{{{ns_uri}}}gIBSUF" if ns_uri else "gIBSUF")
    gibsuf.append(_mk(ns_uri, "vDif", "0.00"))
    gibsuf.append(_mk(ns_uri, "vDevTrib", "0.00"))
    gibsuf.append(_mk(ns_uri, "vIBSUF", _fmt_q2(total_vibsuf)))
    gibs.append(gibsuf)

    gibsmun = etree.Element(f"{{{ns_uri}}}gIBSMun" if ns_uri else "gIBSMun")
    gibsmun.append(_mk(ns_uri, "vDif", "0.00"))
    gibsmun.append(_mk(ns_uri, "vDevTrib", "0.00"))
    gibsmun.append(_mk(ns_uri, "vIBSMun", _fmt_q2(total_vibsmun)))
    gibs.append(gibsmun)

    gibs.append(_mk(ns_uri, "vIBS", _fmt_q2(total_vibs)))
    gibs.append(_mk(ns_uri, "vCredPres", "0.00"))
    gibs.append(_mk(ns_uri, "vCredPresCondSus", "0.00"))
    ibscbstot.append(gibs)

    gcbs = etree.Element(f"{{{ns_uri}}}gCBS" if ns_uri else "gCBS")
    gcbs.append(_mk(ns_uri, "vDif", "0.00"))
    gcbs.append(_mk(ns_uri, "vDevTrib", "0.00"))
    gcbs.append(_mk(ns_uri, "vCBS", _fmt_q2(total_vcbs)))
    gcbs.append(_mk(ns_uri, "vCredPres", "0.00"))
    gcbs.append(_mk(ns_uri, "vCredPresCondSus", "0.00"))
    ibscbstot.append(gcbs)

    _append_or_replace(total, ibscbstot, "IBSCBSTot")

    vnftot = etree.Element(f"{{{ns_uri}}}vNFTot" if ns_uri else "vNFTot")
    vnftot.text = _fmt_q2(total_vbc)
    existentes_vnftot = [
        no for no in total
        if isinstance(no.tag, str) and etree.QName(no).localname == "vNFTot"
    ]
    if existentes_vnftot:
        idx_vnftot = total.index(existentes_vnftot[0])
        for extra in existentes_vnftot:
            total.remove(extra)
        total.insert(idx_vnftot, vnftot)
    else:
        ibscbstot_existente = next(iter(
            no for no in total
            if isinstance(no.tag, str) and etree.QName(no).localname == "IBSCBSTot"
        ), None)
        if ibscbstot_existente is not None:
            idx_insert = total.index(ibscbstot_existente) + 1
        else:
            idx_insert = total.index(icmstot) + 1
        total.insert(idx_insert, vnftot)

    stats["ibscbs_totais_gerados"] = 1


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------

def _resumir_camex_matches(matches: List[Dict], limite: int = 8) -> str:
    partes = []
    for m in matches[:limite]:
        ex = f" Ex {m.get('ex')}" if m.get("ex") else ""
        inicio = m.get("inicio_vigencia") or ""
        fim = m.get("fim_vigencia") or ""
        vig = ""
        if inicio or fim:
            vig = f" ({inicio or 'sem inicio'} a {fim or 'sem termino'})"
        ato = f" - {m.get('ato_legal')}" if m.get("ato_legal") else ""
        partes.append(f"{m.get('lista', '')}{ex}{vig}{ato}")
    if len(matches) > limite:
        partes.append(f"+{len(matches) - limite} lista(s)")
    return " | ".join(partes)


def _resumir_ttd409_matches(matches: List[Dict]) -> str:
    partes = []
    for m in matches:
        partes.append(f"Item {m.get('item')} - {m.get('descricao_legal', '')}")
    return " | ".join(partes)


def _verificar_ttd409(
    root: etree._Element,
    stats: Dict,
    nome_arquivo: str,
    ttd409_lookup: Callable[[str], List[Dict]],
) -> None:
    for det in root.xpath(".//*[local-name()='det']"):
        prod = next(iter(det.xpath("./*[local-name()='prod']")), None)
        if prod is None:
            continue

        ncm = _get_text(prod, "NCM") or ""
        if not ncm:
            continue

        matches = ttd409_lookup(ncm)
        if not matches:
            continue

        stats["ttd409_bloqueios"] += 1
        observacoes = " | ".join(m.get("observacao", "") for m in matches if m.get("observacao"))
        stats["ttd409_itens_detalhado"].append({
            "Gravidade": "GRAVE",
            "Arquivo XML": nome_arquivo,
            "nItem": det.get("nItem", ""),
            "SKU": _get_text(prod, "cProd") or "",
            "NCM": ncm,
            "Descricao": _get_text(prod, "xProd") or "",
            "Regra TTD409": _resumir_ttd409_matches(matches),
            "Observacao": observacoes,
            "Acao recomendada": "Nao incluir no TTD409 sem revisar descricao legal, NCM e excecoes aplicaveis.",
        })


def _verificar_camex(
    root: etree._Element,
    stats: Dict,
    nome_arquivo: str,
    camex_lookup: Callable[[str, str], List[Dict]],
) -> None:
    for det in root.xpath(".//*[local-name()='det']"):
        prod = next(iter(det.xpath("./*[local-name()='prod']")), None)
        if prod is None:
            continue

        ncm = _get_text(prod, "NCM") or ""
        if not ncm:
            continue

        extipi = _get_text(prod, "EXTIPI") or ""
        matches = camex_lookup(ncm, extipi)
        if not matches:
            continue

        stats["camex_alertas"] += 1
        stats["camex_itens_detalhado"].append({
            "Arquivo XML": nome_arquivo,
            "nItem": det.get("nItem", ""),
            "SKU": _get_text(prod, "cProd") or "",
            "NCM": ncm,
            "EXTIPI": extipi,
            "Descricao": _get_text(prod, "xProd") or "",
            "Qtd. listas": len(matches),
            "Listas CAMEX": _resumir_camex_matches(matches),
        })


def auditar_ttd409_xml(
    xml_bytes: bytes,
    nome_arquivo: str,
    ttd409_lookup: Callable[[str], List[Dict]],
) -> Tuple[List[Dict], Dict]:
    """
    Audita um XML apenas contra a base TTD409, sem alterar o arquivo.
    Retorna (itens, stats), incluindo itens bloqueados e itens sem match.
    """
    stats = {
        "arquivo": nome_arquivo,
        "itens_analisados": 0,
        "bloqueios_ttd409": 0,
        "itens_sem_ncm": 0,
        "erros": [],
    }

    try:
        parser = etree.XMLParser(remove_blank_text=False, recover=True, huge_tree=True)
        tree = etree.parse(io.BytesIO(xml_bytes), parser)
        root = tree.getroot()
    except etree.XMLSyntaxError as e:
        stats["erros"].append(f"XML invalido: {e}")
        return [], stats
    except Exception as e:
        stats["erros"].append(f"Erro ao ler o XML: {e}")
        return [], stats

    infnfe = next(iter(root.xpath(".//*[local-name()='infNFe']")), None)
    chave_nfe = ""
    if infnfe is not None:
        chave_nfe = (infnfe.get("Id") or "").replace("NFe", "", 1)

    emissao = (
        _first(root.xpath(".//*[local-name()='ide']/*[local-name()='dhEmi']/text()"))
        or _first(root.xpath(".//*[local-name()='ide']/*[local-name()='dEmi']/text()"))
        or ""
    )

    itens = []
    for det in root.xpath(".//*[local-name()='det']"):
        prod = next(iter(det.xpath("./*[local-name()='prod']")), None)
        if prod is None:
            continue

        stats["itens_analisados"] += 1
        ncm = _get_text(prod, "NCM") or ""
        matches = ttd409_lookup(ncm) if ncm else []

        if not ncm:
            stats["itens_sem_ncm"] += 1
            gravidade = "ATENCAO"
            status = "SEM NCM NO XML"
            regra = ""
            observacao = "Item sem NCM no XML. Conferir manualmente antes de usar no TTD409."
            acao = "Revisar cadastro/NCM do item antes da emissao."
        elif matches:
            stats["bloqueios_ttd409"] += 1
            gravidade = "GRAVE"
            status = "BLOQUEADO / REVISAR TTD409"
            regra = _resumir_ttd409_matches(matches)
            observacao = " | ".join(m.get("observacao", "") for m in matches if m.get("observacao"))
            acao = "Nao incluir no TTD409 sem revisar descricao legal, NCM e excecoes aplicaveis."
        else:
            gravidade = ""
            status = "NAO ENCONTRADO NA LISTA TTD409"
            regra = ""
            observacao = "Sem correspondencia na base TTD409 carregada."
            acao = "Manter conferencia fiscal normal."

        itens.append({
            "Gravidade": gravidade,
            "Status TTD409": status,
            "Arquivo XML": nome_arquivo,
            "Chave NF-e": chave_nfe,
            "Emissao": emissao[:10],
            "nItem": det.get("nItem", ""),
            "SKU": _get_text(prod, "cProd") or "",
            "NCM": ncm,
            "CFOP": _get_text(prod, "CFOP") or "",
            "Descricao": _get_text(prod, "xProd") or "",
            "Quantidade": _get_text(prod, "qCom") or "",
            "Valor": _get_text(prod, "vProd") or "",
            "Regra TTD409": regra,
            "Observacao": observacao,
            "Acao recomendada": acao,
        })

    return itens, stats


def _append_dict_sheet(wb: Workbook, title: str, rows: List[Dict], headers: List[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(idx)]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 55)


def gerar_relatorio_ttd409(auditorias: List[Dict]) -> bytes:
    """Gera Excel consolidado da auditoria TTD409 de varios XMLs."""
    todos_itens = []
    resumo = []
    for auditoria in auditorias:
        itens = auditoria.get("itens", [])
        stats = auditoria.get("stats", {})
        todos_itens.extend(itens)
        resumo.append({
            "Arquivo XML": stats.get("arquivo", auditoria.get("nome_original", "")),
            "Itens analisados": stats.get("itens_analisados", 0),
            "Bloqueios TTD409": stats.get("bloqueios_ttd409", 0),
            "Itens sem NCM": stats.get("itens_sem_ncm", 0),
            "Erros": " | ".join(stats.get("erros", [])),
        })

    bloqueios = [r for r in todos_itens if r.get("Gravidade") == "GRAVE"]

    headers_itens = [
        "Gravidade", "Status TTD409", "Arquivo XML", "Chave NF-e", "Emissao",
        "nItem", "SKU", "NCM", "CFOP", "Descricao", "Quantidade", "Valor",
        "Regra TTD409", "Observacao", "Acao recomendada",
    ]
    headers_resumo = [
        "Arquivo XML", "Itens analisados", "Bloqueios TTD409", "Itens sem NCM", "Erros",
    ]

    wb = Workbook()
    wb.remove(wb.active)
    _append_dict_sheet(wb, "Bloqueios TTD409", bloqueios, headers_itens)
    _append_dict_sheet(wb, "Todos os itens", todos_itens, headers_itens)
    _append_dict_sheet(wb, "Resumo", resumo, headers_resumo)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def processar_xml(
    xml_bytes: bytes,
    mapa_ean: Dict[str, str],
    nome_arquivo: str,
    mapa_fiscal: Optional[Dict[str, Dict]] = None,
    camex_lookup: Optional[Callable[[str, str], List[Dict]]] = None,
    ttd409_lookup: Optional[Callable[[str], List[Dict]]] = None,
) -> Tuple[Optional[bytes], Dict]:
    """
    Processa um único XML NF-e aplicando todas as regras.
    Retorna (bytes_processado, stats).
    """
    stats = _novo_stats()

    try:
        parser = etree.XMLParser(remove_blank_text=False, recover=True, huge_tree=True)
        tree = etree.parse(io.BytesIO(xml_bytes), parser)
        root = tree.getroot()
    except etree.XMLSyntaxError as e:
        stats["erros"].append(f"XML inválido: {e}")
        return None, stats
    except Exception as e:
        stats["erros"].append(f"Erro ao ler o XML: {e}")
        return None, stats

    # 0. TTD409 / Decreto SC 2.128/2009 - alerta grave, nao altera o XML
    if ttd409_lookup is not None:
        try:
            _verificar_ttd409(root, stats, nome_arquivo, ttd409_lookup)
        except Exception as e:
            stats["avisos"].append(f"Falha ao verificar TTD409 (ignorado): {e}")

    # 0.1 CAMEX / Gecex - apenas alerta, nao altera o XML
    if camex_lookup is not None:
        try:
            _verificar_camex(root, stats, nome_arquivo, camex_lookup)
        except Exception as e:
            stats["avisos"].append(f"Falha ao verificar CAMEX (ignorado): {e}")

    # 1. EAN
    try:
        _aplicar_ean(root, stats, mapa_ean, nome_arquivo)
    except Exception as e:
        stats["avisos"].append(f"Falha ao aplicar EAN (ignorado): {e}")

    # 2. cFabricante
    try:
        _inserir_cfabricante(root, stats)
    except Exception as e:
        stats["avisos"].append(f"Falha ao inserir cFabricante (ignorado): {e}")

    # 3. CFOP
    try:
        _ajustar_cfop(root, stats)
    except Exception as e:
        stats["avisos"].append(f"Falha ao ajustar CFOP (ignorado): {e}")

    # 4. ICMS zerado
    try:
        _zerar_icms(root, stats)
    except Exception as e:
        stats["avisos"].append(f"Falha ao zerar ICMS (ignorado): {e}")

    # 5. IPI
    try:
        _ajustar_ipi(root, stats)
    except Exception as e:
        stats["avisos"].append(f"Falha ao ajustar IPI (ignorado): {e}")

    # 6. PIS/COFINS
    try:
        _ajustar_pis_cofins(root, stats)
    except Exception as e:
        stats["avisos"].append(f"Falha ao ajustar PIS/COFINS (ignorado): {e}")

    # 7. IBS/CBS
    try:
        _adicionar_ibscbs(root, stats, mapa_fiscal)
    except Exception as e:
        stats["avisos"].append(f"Falha ao incluir IBS/CBS (ignorado): {e}")

    # Serializar
    try:
        encoding = tree.docinfo.encoding or "UTF-8"
        out = io.BytesIO()
        tree.write(out, encoding=encoding, xml_declaration=True, pretty_print=False)
        return out.getvalue(), stats
    except Exception as e:
        stats["erros"].append(f"Erro ao serializar XML: {e}")
        return None, stats


# ---------------------------------------------------------------------------
# Relatório Excel — SKUs sem EAN
# ---------------------------------------------------------------------------

def gerar_relatorio_faltantes(resultados: List[Dict]) -> bytes:
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumo"
    ws_res.append(["SKU", "Qtd. itens sem EAN"])

    contagem: Dict[str, int] = {}
    detalhes: List[Dict] = []
    for r in resultados:
        if r["stats"].get("erros"):
            continue
        for item in r["stats"].get("faltantes_detalhado", []):
            sku = item["SKU"]
            contagem[sku] = contagem.get(sku, 0) + 1
            detalhes.append(item)

    for sku, qtd in sorted(contagem.items(), key=lambda x: (-x[1], x[0])):
        ws_res.append([sku, qtd])

    ws_det = wb.create_sheet("Detalhado")
    ws_det.append(["Arquivo XML", "nItem", "SKU", "Descrição", "cEAN no XML", "cEANTrib no XML"])
    for row in detalhes:
        ws_det.append([
            row.get("Arquivo XML", ""),
            row.get("nItem", ""),
            row.get("SKU", ""),
            row.get("Descrição", ""),
            row.get("cEAN no XML", ""),
            row.get("cEANTrib no XML", ""),
        ])

    for ws in (ws_res, ws_det):
        for col in ws.columns:
            maxlen = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(maxlen + 3, 60)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
